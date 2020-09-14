# -*- coding: UTF-8 -*-
# @Time     :2020/9/12 16:32
# @Author   :raoyu
# @Email    :2458757210@qq.com
# @Phone    :18893703014
import openpyxl
import os
from Common.handle_path import DataPath


class Handle_Excel:
    """处理Excel文件数据"""

    def __init__(self, filename, sheetname):
        """
        :param filename: 测试数据文件路径
        :param sheetname: 工作薄的sheet页名称
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建sheet页对象
        sh = wb[self.sheetname]
        # 按行读取excel文件数据
        excel_datas = list(sh.rows)
        # 创建空列表存放测试数据
        case_datas = []
        # 创建空列表存放标题数据
        titles = []
        # 循环遍历出标题的值
        for title in excel_datas[0]:
            titles.append(title.value)

        for datas in excel_datas[1:]:
            values = []
            for item in datas:
                values.append(item.value)
            # 将标题和测试数据值打包为字典
            case_data = dict(zip(titles, values))
            case_datas.append(case_data)
        return case_datas

    def write_excel(self, row, column, value):
        # 创建工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 创建sheet页对象
        sh = wb[self.sheetname]
        # 写入数值到excel
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)


if __name__ == '__main__':
    DataPath = os.path.join(DataPath, 'gs_testcase.xlsx')
    excel = Handle_Excel(DataPath, 'login')
    data = excel.read_excel()
    print(data)
